import sys
import os
import json
import io
import hashlib
import tempfile
import traceback
from typing import Dict, Any, Optional, Tuple, List

from PySide6.QtCore import Qt, Signal, QObject, QThread, QTimer, QRunnable, QThreadPool, Slot
from PySide6.QtGui import (
    QIcon, QPixmap, QPen, QColor, QImage, QPainter,
    QFontDatabase, QTextCursor
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QComboBox, QPushButton, QLineEdit, QMessageBox, QFileDialog,
    QDialog, QFormLayout, QDoubleSpinBox,
    QTextEdit, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem, QGraphicsRectItem,
    QCheckBox, QFrame, QSizePolicy, QSpinBox, QProgressDialog, QStyle, QScrollArea
)

from openpyxl import load_workbook
import win32print

PIL_AVAILABLE = True
BARCODE_AVAILABLE = True
try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:
    PIL_AVAILABLE = False

try:
    from barcode import get_barcode_class
    from barcode.writer import ImageWriter
except Exception:
    BARCODE_AVAILABLE = False


APP_TITLE = "Label Print"
SETTINGS_FILE = "settings.json"
LAYOUT_FILE = "layout.json"
DEFAULT_DPI = 203


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller bundle."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# ----------------------------
# Crash guard (best-effort)
# ----------------------------
def _install_global_exception_hook():
    """
    Best-effort safety net: prevents silent hard-crashes from uncaught exceptions
    by showing a message box (when possible) and printing traceback.
    """
    def _hook(exctype, value, tb):
        try:
            msg = "".join(traceback.format_exception(exctype, value, tb))
            print(msg, file=sys.stderr)

            app = QApplication.instance()
            if app is not None:
                QMessageBox.critical(
                    None,
                    "Unexpected Error",
                    "The application hit an unexpected error.\n\n"
                    "Details (also printed to console):\n\n"
                    f"{msg[-4000:]}"
                )
        except Exception:
            # Never let the hook crash
            pass

    sys.excepthook = _hook


# ----------------------------
# Helpers
# ----------------------------
def mm_to_inches(mm: float) -> float:
    return mm / 25.4


def inches_to_mm(inches: float) -> float:
    return inches * 25.4


def mm_to_dots(mm: float, dpi: int) -> int:
    return int(round((mm / 25.4) * dpi))


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def ensure_dir_file(path: str) -> None:
    d = os.path.dirname(os.path.abspath(path))
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)


def normalize_price(text: str) -> str:
    t = (text or "").strip().replace("$", "")
    if not t:
        return "$0.00"
    try:
        v = float(t)
        return f"${v:.2f}"
    except Exception:
        return f"${t}"


def is_numeric_code(text: str) -> bool:
    t = (text or "").strip()
    return bool(t) and t.isdigit()


def parse_csv_line(line: str) -> Optional[Tuple[str, str, str]]:
    if not line:
        return None
    raw = line.strip()
    if not raw:
        return None
    parts = [p.strip() for p in raw.split(",")]

    # Need at least: description, price
    if len(parts) < 2:
        return None

    price = parts[-1].strip()

    # description, price
    if len(parts) == 2:
        code = ""
        desc = parts[0].strip()
        return code, desc, price

    # 3+ parts: detect if first token is numeric code
    first = (parts[0] or "").strip()

    if first == "":
        code = ""
        desc = ", ".join(parts[1:-1]).strip()
        return code, desc, price

    if is_numeric_code(first):
        code = first
        desc = ", ".join(parts[1:-1]).strip()
        return code, desc, price

    # First token is not numeric -> no code, treat everything except last as description
    code = ""
    desc = ", ".join(parts[:-1]).strip()
    return code, desc, price


def sha1_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()


def rects_intersect(a: Tuple[int, int, int, int], b: Tuple[int, int, int, int]) -> bool:
    ax, ay, aw, ah = a
    bx, by, bw, bh = b
    if aw <= 0 or ah <= 0 or bw <= 0 or bh <= 0:
        return False
    return not (ax + aw <= bx or bx + bw <= ax or ay + ah <= by or by + bh <= ay)


# ----------------------------
# Settings store
# ----------------------------
class SettingsStore:
    def __init__(self, path: str):
        self.path = path
        self.data: Dict[str, Any] = {}

    def load(self) -> None:
        if os.path.exists(self.path):
            with open(self.path, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        else:
            self.data = {}

        self.data.setdefault("dpi", DEFAULT_DPI)
        self.data.setdefault("unit", "inches")
        self.data.setdefault("label_sizes_mm", [[76.2, 25.4], [50, 25], [60, 40], [70, 40], [100, 50]])
        self.data.setdefault("selected_label_size_mm", [76.2, 25.4])
        self.data.setdefault("selected_printer", "")
        self.data.setdefault("barcode_type", "Code128")
        self.data.setdefault("last_excel_dir", "")
        self.data.setdefault("lock_font_family", True)
        self.data.setdefault("company_name", "")

    def save(self) -> None:
        ensure_dir_file(self.path)
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2)

    @property
    def dpi(self) -> int:
        return int(self.data.get("dpi", DEFAULT_DPI))

    @property
    def unit(self) -> str:
        return str(self.data.get("unit", "inches"))

    @property
    def label_sizes_mm(self) -> List[List[float]]:
        return list(self.data.get("label_sizes_mm", []))

    @property
    def selected_label_size_mm(self) -> Tuple[float, float]:
        w, h = self.data.get("selected_label_size_mm", [76.2, 25.4])
        return float(w), float(h)

    @property
    def selected_printer(self) -> str:
        return str(self.data.get("selected_printer", ""))

    @property
    def barcode_type(self) -> str:
        return str(self.data.get("barcode_type", "Code128"))


# ----------------------------
# Layout store (relative coords)
# ----------------------------
class LayoutStore:
    def __init__(self, path: str):
        self.path = path
        self.layouts: Dict[str, Any] = {}
        self.load()

    def load(self) -> None:
        self.layouts = {}
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    self.layouts = json.load(f)
                return
            except Exception:
                self.layouts = {}

        bundled = resource_path(LAYOUT_FILE)
        try:
            if os.path.exists(bundled) and os.path.abspath(bundled) != os.path.abspath(self.path):
                with open(bundled, "r", encoding="utf-8") as f:
                    self.layouts = json.load(f)
                try:
                    self.save()
                except Exception:
                    pass
        except Exception:
            self.layouts = {}

    def save(self) -> None:
        ensure_dir_file(self.path)
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.layouts, f, indent=2)

    @staticmethod
    def key_for_size_mm(w_mm: float, h_mm: float) -> str:
        return f"{round(w_mm, 3)}x{round(h_mm, 3)}"

    def get_layout(self, w_mm: float, h_mm: float) -> Dict[str, Any]:
        key = self.key_for_size_mm(w_mm, h_mm)
        if key in self.layouts:
            layout = self.layouts[key]
        else:
            layout = {
                "auto_fit": True,
                "font_family": "Arial",
                "elements": {
                    "description": {"x": 0.50, "y": 0.18, "font": 18, "bold": True, "align": "center"},
                    "price": {"x": 0.50, "y": 0.48, "font": 26, "bold": True, "align": "center"},
                    "barcode": {"x": 0.07, "y": 0.68, "module_w": 2.0, "bar_h_frac": 0.18, "width_frac": 0.78, "quiet_zone": 3.0},
                    "barcode_text": {"x": 0.10, "y": 0.88, "font": 12, "bold": True, "align": "left"},
                    "company": {"x": 0.92, "y": 0.88, "font": 14, "bold": True, "align": "right"}
                }
            }
            self.layouts[key] = layout
            self.save()

        els = layout.get("elements", {})
        bc = els.get("barcode", {})
        if "width_frac" not in bc:
            bc["width_frac"] = 0.78
        if "quiet_zone" not in bc:
            bc["quiet_zone"] = 3.0
        els["barcode"] = bc

        for k in ("description", "price", "barcode_text", "company"):
            if k in els:
                els[k]["align"] = "center"

        layout["elements"] = els
        layout.setdefault("font_family", "Arial")
        return layout

    def set_layout(self, w_mm: float, h_mm: float, layout: Dict[str, Any]) -> None:
        key = self.key_for_size_mm(w_mm, h_mm)
        self.layouts[key] = layout
        self.save()


# ----------------------------
# ZPL Bitmap helpers (GFA)
# ----------------------------
def pil_to_1bpp_bytes(pil_l: "Image.Image") -> Tuple[bytes, int, int, int]:
    if pil_l.mode != "L":
        pil_l = pil_l.convert("L")

    w, h = pil_l.size
    bw = pil_l.point(lambda p: 0 if p < 128 else 255, mode="L")
    bpr = (w + 7) // 8
    out = bytearray(bpr * h)
    px = bw.load()

    idx = 0
    for y in range(h):
        for bx in range(bpr):
            byte = 0
            for bit in range(8):
                x = bx * 8 + bit
                if x >= w:
                    v_black = 0
                else:
                    v_black = 1 if px[x, y] == 0 else 0
                byte = (byte << 1) | v_black
            out[idx] = byte
            idx += 1

    return bytes(out), w, h, bpr


def zpl_gfa_from_1bpp(data: bytes, width: int, height: int, bytes_per_row: int) -> str:
    total = len(data)
    hexdata = data.hex().upper()
    return f"^GFA,{total},{total},{bytes_per_row},{hexdata}"


def zpl_for_bitmap_label(w_dots: int, h_dots: int, gfa_cmd: str) -> str:
    return "\n".join([
        "^XA",
        f"^PW{w_dots}",
        f"^LL{h_dots}",
        "^LH0,0",
        "^FO0,0",
        gfa_cmd,
        "^FS",
        "^XZ"
    ])


# ----------------------------
# Bitmap Label Renderer
# ----------------------------
class BitmapLabelRenderer:
    def __init__(self):
        self._barcode_cache: Dict[str, "Image.Image"] = {}
        self._font_cache: Dict[str, "ImageFont.FreeTypeFont"] = {}
        self._font_path_cache: Dict[Tuple[str, bool], Optional[str]] = {}

    def _find_font_path(self, family: str, bold: bool) -> Optional[str]:
        fam = (family or "Arial").strip().lower()
        cache_key = (fam, bool(bold))
        if cache_key in self._font_path_cache:
            return self._font_path_cache[cache_key]

        font_map = {
            "arial": (["arial.ttf"], ["arialbd.ttf"]),
            "arial black": (["ariblk.ttf"], ["ariblk.ttf"]),
            "bahnschrift": (["bahnschrift.ttf"], ["bahnschrift.ttf"]),
            "calibri": (["calibri.ttf"], ["calibrib.ttf"]),
            "calibri light": (["calibril.ttf"], ["calibril.ttf"]),
            "cambria": (["cambria.ttc"], ["cambriab.ttf"]),
            "candara": (["candara.ttf"], ["candarab.ttf"]),
            "cascadia code": (["cascadiacode.ttf"], ["cascadiacode.ttf"]),
            "cascadia mono": (["cascadiamono.ttf"], ["cascadiamono.ttf"]),
            "century gothic": (["gothic.ttf"], ["gothicb.ttf"]),
            "comic sans ms": (["comic.ttf"], ["comicbd.ttf"]),
            "consolas": (["consola.ttf"], ["consolab.ttf"]),
            "constantia": (["constan.ttf"], ["constanb.ttf"]),
            "corbel": (["corbel.ttf"], ["corbelb.ttf"]),
            "courier new": (["cour.ttf"], ["courbd.ttf"]),
            "gadugi": (["gadugi.ttf"], ["gadugib.ttf"]),
            "segoe ui": (["segoeui.ttf"], ["segoeuib.ttf"]),
            "segoe ui black": (["seguibl.ttf"], ["seguibl.ttf"]),
            "segoe ui light": (["segoeuil.ttf"], ["segoeuil.ttf"]),
            "segoe ui semibold": (["seguisb.ttf"], ["seguisb.ttf"]),
            "segoe ui semilight": (["segoeuisl.ttf"], ["segoeuisl.ttf"]),
            "verdana": (["verdana.ttf"], ["verdanab.ttf"]),
            "tahoma": (["tahoma.ttf"], ["tahomabd.ttf", "tahoma.ttf"]),
            "times new roman": (["times.ttf"], ["timesbd.ttf"]),
            "georgia": (["georgia.ttf"], ["georgiab.ttf"]),
            "trebuchet ms": (["trebuc.ttf"], ["trebucbd.ttf"]),
            "impact": (["impact.ttf"], ["impact.ttf"]),
            "ink free": (["inkfree.ttf"], ["inkfree.ttf"]),
            "lucida console": (["lucon.ttf"], ["lucon.ttf"]),
            "palatino linotype": (["pala.ttf"], ["palab.ttf"]),
        }

        win_dir = os.environ.get("WINDIR", r"C:\Windows")
        fonts_dir = os.path.join(win_dir, "Fonts")

        if fam in font_map:
            regular, bolds = font_map[fam]
            candidates = (bolds if bold else regular)
            for c in candidates:
                p = os.path.join(fonts_dir, c)
                if os.path.exists(p):
                    self._font_path_cache[cache_key] = p
                    return p

        target = fam.replace(" ", "")
        try:
            files = os.listdir(fonts_dir)
        except Exception:
            files = []

        bold_tokens = ("bd", "bold", "black", "semibold")
        exts = (".ttf", ".otf", ".ttc")

        matches = []
        for fn in files:
            low = fn.lower()
            if not low.endswith(exts):
                continue
            low_nospace = low.replace(" ", "")
            if target and target in low_nospace:
                matches.append(fn)

        if bold and matches:
            bold_matches = [m for m in matches if any(tok in m.lower() for tok in bold_tokens)]
            if bold_matches:
                matches = bold_matches

        if matches:
            p = os.path.join(fonts_dir, matches[0])
            if os.path.exists(p):
                self._font_path_cache[cache_key] = p
                return p

        self._font_path_cache[cache_key] = None
        return None

    def _get_font(self, family: str, size_px: int, bold: bool) -> "ImageFont.ImageFont":
        family_safe = (family or "Arial").strip() or "Arial"
        key = f"{family_safe}|{size_px}|{1 if bold else 0}"
        if key in self._font_cache:
            return self._font_cache[key]

        path = self._find_font_path(family_safe, bold)
        try:
            if path and PIL_AVAILABLE:
                f = ImageFont.truetype(path, size_px)
            else:
                if family_safe.lower() != "arial":
                    path2 = self._find_font_path("Arial", bold)
                    if path2 and PIL_AVAILABLE:
                        f = ImageFont.truetype(path2, size_px)
                    else:
                        f = ImageFont.load_default()
                else:
                    f = ImageFont.load_default()
        except Exception:
            f = ImageFont.load_default()

        self._font_cache[key] = f
        return f

    def _text_bbox(self, draw: "ImageDraw.ImageDraw", text: str, font: "ImageFont.ImageFont") -> Tuple[int, int]:
        if not text:
            return (0, 0)
        try:
            b = draw.textbbox((0, 0), text, font=font)
            return (b[2] - b[0], b[3] - b[1])
        except Exception:
            return (int(len(text) * 8), 14)

    def _truncate_to_width(self, draw: "ImageDraw.ImageDraw", text: str, font: "ImageFont.ImageFont", max_w: int) -> str:
        if not text:
            return text
        w, _ = self._text_bbox(draw, text, font)
        if w <= max_w:
            return text
        ell = "..."
        lo, hi = 0, len(text)
        best = ""
        while lo <= hi:
            mid = (lo + hi) // 2
            candidate = text[:mid].rstrip() + ell
            cw, _ = self._text_bbox(draw, candidate, font)
            if cw <= max_w:
                best = candidate
                lo = mid + 1
            else:
                hi = mid - 1
        return best if best else ell

    def _anchor_pos(self, w: int, h: int, x_frac: float, y_frac: float, bw: int, bh: int, align: str) -> Tuple[int, int]:
        ax = int(round(x_frac * w))
        ay = int(round(y_frac * h))
        y = ay - bh // 2
        if align == "center":
            x = ax - bw // 2
        elif align == "right":
            x = ax - bw
        else:
            x = ax
        return x, y

    def _barcode_key(self, code: str, btype: str, module_w_dots: int, bar_h_dots: int, target_w: int, quiet_zone: float) -> str:
        s = f"{btype}|{code}|mw{module_w_dots}|h{bar_h_dots}|w{target_w}|qz{quiet_zone:.2f}"
        return sha1_text(s)

    def _render_barcode(
        self,
        code: str,
        btype: str,
        module_w_dots: int,
        bar_h_dots: int,
        target_w: int,
        quiet_zone: float
    ) -> "Image.Image":
        img = Image.new("L", (max(1, target_w), max(1, bar_h_dots)), 255)
        if not PIL_AVAILABLE or not BARCODE_AVAILABLE:
            return img

        code = (code or "").strip() or "0"
        bt = (btype or "Code128").strip().lower()
        if bt not in ("code128", "code39"):
            bt = "code128"

        qz = float(clamp(float(quiet_zone), 0.0, 20.0))

        key = self._barcode_key(code, bt, int(module_w_dots), int(bar_h_dots), int(target_w), qz)
        if key in self._barcode_cache:
            return self._barcode_cache[key].copy()

        try:
            klass = get_barcode_class(bt)
            b = klass(code, writer=ImageWriter())
            fp = io.BytesIO()

            module_width = clamp(0.12 + 0.06 * float(module_w_dots), 0.14, 0.80)
            module_height = max(10, int(bar_h_dots * 0.95))

            b.write(fp, options={
                "module_width": module_width,
                "module_height": module_height,
                "quiet_zone": qz,
                "write_text": False,
                "background": "white",
                "foreground": "black",
                "dpi": 300,
            })
            fp.seek(0)
            pil = Image.open(fp).convert("L")
            pil = pil.resize((max(1, target_w), max(1, bar_h_dots)), resample=Image.Resampling.NEAREST)

            self._barcode_cache[key] = pil.copy()
            return pil
        except Exception:
            return img

    def render(
        self,
        dpi: int,
        w_mm: float,
        h_mm: float,
        layout: Dict[str, Any],
        barcode_type: str,
        company: str,
        item_code: str,
        desc: str,
        price: str
    ) -> Tuple["Image.Image", Dict[str, Tuple[int, int, int, int]], List[str]]:

        if not PIL_AVAILABLE:
            w_dots = mm_to_dots(w_mm, dpi)
            h_dots = mm_to_dots(h_mm, dpi)
            img = Image.new("L", (w_dots, h_dots), 255)
            return img, {}, ["PIL missing"]

        w_dots = mm_to_dots(w_mm, dpi)
        h_dots = mm_to_dots(h_mm, dpi)

        img = Image.new("L", (w_dots, h_dots), 255)
        draw = ImageDraw.Draw(img)

        elements = layout.get("elements", {})
        auto_fit = bool(layout.get("auto_fit", True))
        font_family = str(layout.get("font_family", "Arial")) or "Arial"

        pad_x = int(round(w_dots * 0.04))
        pad_y = int(round(h_dots * 0.04))

        bboxes: Dict[str, Tuple[int, int, int, int]] = {}
        issues: List[str] = []

        item_code_clean = (item_code or "").strip()
        has_code = bool(item_code_clean)

        def place_text(name: str, text: str, spec: Dict[str, Any], max_width: Optional[int] = None):
            if text is None:
                text = ""
            align = "center"
            bold = bool(spec.get("bold", False))
            size_pt = int(spec.get("font", 14))

            size_px = max(8, int(round(size_pt * (dpi / 72.0))))
            font = self._get_font(font_family, size_px, bold)

            text_draw = text.strip()

            if max_width is not None:
                if auto_fit:
                    test_size = size_px
                    while test_size > 8:
                        font = self._get_font(font_family, test_size, bold)
                        tw, th = self._text_bbox(draw, text_draw, font)
                        if tw <= max_width:
                            break
                        test_size -= 1
                    font = self._get_font(font_family, max(8, test_size), bold)

                tw, th = self._text_bbox(draw, text_draw, font)
                if tw > max_width:
                    text_draw = self._truncate_to_width(draw, text_draw, font, max_width)
                    tw, th = self._text_bbox(draw, text_draw, font)
            else:
                tw, th = self._text_bbox(draw, text_draw, font)

            x, y = self._anchor_pos(
                w_dots, h_dots,
                float(spec.get("x", 0.1)),
                float(spec.get("y", 0.1)),
                tw, th,
                align
            )

            # IMPORTANT FIX:
            # Do NOT clamp x/y based on text box width/height.
            # This previously "froze" arrow movement when the text was wide.
            # We allow free positioning; clipping is reported below via issue checks.

            try:
                draw.text((x, y), text_draw, font=font, fill=0)
            except Exception:
                draw.text((x, y), text_draw, font=ImageFont.load_default(), fill=0)

            bboxes[name] = (x, y, tw, th)

        if "description" in elements:
            place_text("Description", desc or "", elements["description"], max_width=int(round(w_dots * 0.92)))

        if "price" in elements:
            place_text("Price", normalize_price(price or ""), elements["price"], max_width=int(round(w_dots * 0.92)))

        if has_code and "barcode" in elements:
            b = elements["barcode"]
            bx = int(round(float(b.get("x", 0.07)) * w_dots))
            by = int(round(float(b.get("y", 0.68)) * h_dots))
            module_w = float(b.get("module_w", 2.0))
            bar_h_frac = float(b.get("bar_h_frac", 0.18))
            width_frac = float(b.get("width_frac", 0.78))
            quiet_zone = float(b.get("quiet_zone", 3.0))

            module_w_dots = max(1, int(round(module_w)))

            desired_h = max(18, int(round(bar_h_frac * h_dots)))
            available_h = max(1, h_dots - by - pad_y)
            bar_h_dots = min(desired_h, available_h)
            if bar_h_dots < desired_h:
                issues.append("Clipping: Barcode")

            desired_w = int(round(clamp(width_frac, 0.10, 0.98) * w_dots))
            available_w = max(1, w_dots - bx - pad_x)
            target_w = min(desired_w, available_w)
            if target_w < desired_w:
                issues.append("Clipping: Barcode")

            bar = self._render_barcode(item_code_clean, barcode_type, module_w_dots, bar_h_dots, target_w, quiet_zone)
            img.paste(bar, (bx, by))
            bboxes["Barcode"] = (bx, by, target_w, bar_h_dots)

        if has_code and "barcode_text" in elements:
            place_text("Barcode Text", item_code_clean, elements["barcode_text"], max_width=int(round(w_dots * 0.92)))

        if "company" in elements:
            place_text("Company Name", (company or "").strip(), elements["company"], max_width=int(round(w_dots * 0.92)))

        # clipping check (this is what you wanted: warn, don't freeze)
        for name, (x, y, bw, bh) in bboxes.items():
            if x < 0 or y < 0 or (x + bw) > w_dots or (y + bh) > h_dots:
                if f"Clipping: {name}" not in issues:
                    issues.append(f"Clipping: {name}")

        overlap_pairs = [
            ("Description", "Price"),
            ("Price", "Barcode"),
            ("Barcode", "Barcode Text"),
            ("Barcode", "Company Name"),
        ]
        for a, b in overlap_pairs:
            if a in bboxes and b in bboxes and rects_intersect(bboxes[a], bboxes[b]):
                issues.append(f"Overlap: {a} + {b}")

        return img, bboxes, issues


# ----------------------------
# Preview render worker
# ----------------------------
class RenderResult:
    def __init__(self, request_key: str, qimage: QImage, bboxes: Dict[str, Tuple[int, int, int, int]], issues: List[str]):
        self.request_key = request_key
        self.qimage = qimage
        self.bboxes = bboxes
        self.issues = issues


class RenderWorker(QObject):
    finished = Signal(object)
    failed = Signal(str)

    def __init__(
        self,
        renderer: BitmapLabelRenderer,
        dpi: int,
        w_mm: float,
        h_mm: float,
        layout: Dict[str, Any],
        barcode_type: str,
        company: str,
        code: str,
        desc: str,
        price: str,
        request_key: str
    ):
        super().__init__()
        self.renderer = renderer
        self.dpi = dpi
        self.w_mm = w_mm
        self.h_mm = h_mm
        self.layout = layout
        self.barcode_type = barcode_type
        self.company = company
        self.code = code
        self.desc = desc
        self.price = price
        self.request_key = request_key

    @Slot()
    def run(self):
        try:
            pil_l, bboxes, issues = self.renderer.render(
                dpi=self.dpi,
                w_mm=self.w_mm,
                h_mm=self.h_mm,
                layout=self.layout,
                barcode_type=self.barcode_type,
                company=self.company,
                item_code=self.code,
                desc=self.desc,
                price=self.price
            )

            if not PIL_AVAILABLE:
                raise RuntimeError("PIL is required for bitmap rendering.")

            w, h = pil_l.size
            raw = pil_l.tobytes("raw", "L")
            qimg = QImage(raw, w, h, w, QImage.Format_Grayscale8).copy()
            self.finished.emit(RenderResult(self.request_key, qimg, bboxes, issues))
        except Exception as e:
            self.failed.emit(str(e))


# ----------------------------
# Clickable GraphicsView (no dragging)
# ----------------------------
class LabelClickView(QGraphicsView):
    def __init__(self, scene: QGraphicsScene, parent=None):
        super().__init__(scene, parent)
        self._click_cb = None

        self.setRenderHint(QPainter.SmoothPixmapTransform, False)
        self.setRenderHint(QPainter.Antialiasing, False)
        self.setViewportUpdateMode(QGraphicsView.MinimalViewportUpdate)
        self.setDragMode(QGraphicsView.NoDrag)

        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

    def set_click_callback(self, cb):
        self._click_cb = cb

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._click_cb:
            sp = self.mapToScene(event.position().toPoint())
            self._click_cb(sp)
        super().mousePressEvent(event)


# ----------------------------
# Preview / Layout
# ----------------------------
class PreviewLayoutDialog(QDialog):
    def __init__(self, parent, settings: SettingsStore, layout_store: LayoutStore, renderer: BitmapLabelRenderer, get_current_row_fn):
        super().__init__(parent)
        self.setWindowTitle("Preview / Layout (Bitmap)")
        self.settings = settings
        self.layout_store = layout_store
        self.renderer = renderer
        self.get_current_row = get_current_row_fn

        self._thread: Optional[QThread] = None
        self._worker: Optional[RenderWorker] = None
        self._active_request_key: str = ""
        self._render_cache: Dict[str, RenderResult] = {}
        self._render_in_flight = False
        self._rerender_queued = False

        self._closing = False
        self._first_show_done = False

        self.w_mm, self.h_mm = self.settings.selected_label_size_mm
        self._loaded_size = (self.w_mm, self.h_mm)
        self.layout_data = json.loads(json.dumps(self.layout_store.get_layout(self.w_mm, self.h_mm)))

        self.selected_element: Optional[str] = None
        self._last_bboxes: Dict[str, Tuple[int, int, int, int]] = {}

        self._pending_timer = QTimer(self)
        self._pending_timer.setSingleShot(True)
        self._pending_timer.timeout.connect(self._start_render)

        self.setStyleSheet("""
            QDialog { background: #f3f9ff; }
            QLabel { color: #0f2b46; font-size: 13px; }
            QLineEdit, QComboBox, QTextEdit, QSpinBox, QDoubleSpinBox {
                background: white;
                color: #0f2b46;
                border: 1px solid #cfe6ff;
                border-radius: 10px;
                padding: 6px;
                font-size: 13px;
            }

            QSpinBox, QDoubleSpinBox { padding-right: 26px; }
            QSpinBox::up-button, QDoubleSpinBox::up-button {
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 18px;
                border-left: 1px solid #cfe6ff;
                border-bottom: 1px solid #cfe6ff;
                border-top-right-radius: 10px;
            }
            QSpinBox::down-button, QDoubleSpinBox::down-button {
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 18px;
                border-left: 1px solid #cfe6ff;
                border-bottom-right-radius: 10px;
            }

            QCheckBox { color: #0f2b46; font-weight: 900; font-size: 13px; }
            QCheckBox::indicator { width: 18px; height: 18px; }
            QPushButton {
                background: #2d7ef7;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px 14px;
                font-weight: 900;
            }
            QPushButton:hover { background: #236be0; }
            QPushButton:pressed { background: #1c58ba; }

            QMessageBox { background: white; }
            QMessageBox QLabel { color: #0f2b46; font-size: 14px; }
            QMessageBox QPushButton { min-width: 110px; }

            QScrollArea { background: white; border: 1px solid #cfe6ff; border-radius: 14px; }
            QScrollArea QWidget { background: white; }
        """)

        root = QHBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(12)

        inspector_widget = QWidget()
        left = QVBoxLayout(inspector_widget)
        left.setSpacing(8)
        left.setContentsMargins(12, 12, 12, 12)

        def section_title(t: str) -> QLabel:
            lb = QLabel(t)
            lb.setStyleSheet("font-weight:900; font-size:16px; color:#0f2b46;")
            return lb

        def small_label(t: str) -> QLabel:
            lb = QLabel(t)
            lb.setStyleSheet("font-weight:900; font-size:13px; color:#0f2b46;")
            return lb

        left.addWidget(section_title("Inspector"))

        self.lbl_selected = QLabel("Selected: (click on label)")
        self.lbl_selected.setStyleSheet("font-weight:900; font-size:13px; color:#0f2b46;")
        left.addWidget(self.lbl_selected)

        self.lbl_issues = QLabel("No issues ✓")
        self.lbl_issues.setWordWrap(True)
        self.lbl_issues.setStyleSheet("font-weight:900; font-size:13px; color:#118a3b;")
        left.addWidget(self.lbl_issues)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color:#cfe6ff;")
        left.addWidget(sep)

        self.chk_autofit = QCheckBox("Auto-fit text to label width")
        self.chk_autofit.setChecked(bool(self.layout_data.get("auto_fit", True)))
        self.chk_autofit.stateChanged.connect(self._autofit_changed)
        left.addWidget(self.chk_autofit)

        left.addWidget(section_title("Theme font"))
        self.font_combo = QComboBox()

        # remove fonts requested + only show fonts that map to a real Windows font file
        banned_fonts = {"MT Extra", "Marlett", "Symbol", "Webdings"}
        common = [
            "Arial",
            "Arial Black",
            "Bahnschrift",
            "Calibri",
            "Calibri Light",
            "Cambria",
            "Candara",
            "Cascadia Code",
            "Cascadia Mono",
            "Century Gothic",
            "Comic Sans MS",
            "Consolas",
            "Constantia",
            "Corbel",
            "Courier New",
            "Gadugi",
            "Impact",
            "Ink Free",
            "Lucida Console",
            "Palatino Linotype",
            "Segoe UI",
            "Segoe UI Black",
            "Segoe UI Light",
            "Segoe UI Semibold",
            "Segoe UI Semilight",
            "Tahoma",
            "Times New Roman",
            "Georgia",
            "Trebuchet MS",
            "Verdana",
        ]

        families = list(QFontDatabase.families())
        seen = set()
        working: List[str] = []

        def add_if_working(name: str):
            n = (name or "").strip()
            if not n or n in seen:
                return
            if n in banned_fonts:
                return
            seen.add(n)
            try:
                if self.renderer._find_font_path(n, False) or self.renderer._find_font_path(n, True):
                    working.append(n)
            except Exception:
                pass

        for f in common:
            add_if_working(f)
        for f in sorted(families):
            add_if_working(f)

        if not working:
            working = ["Arial"]
        else:
            working = sorted(set(working), key=lambda s: s.lower())

        self.font_combo.addItems(working)

        cur_ff = str(self.layout_data.get("font_family", "Arial")) or "Arial"
        if cur_ff in banned_fonts or cur_ff not in working:
            cur_ff = "Arial"
            self.layout_data["font_family"] = "Arial"
        self.font_combo.setCurrentText(cur_ff)

        self.font_combo.currentTextChanged.connect(self._font_changed)
        left.addWidget(self.font_combo)

        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet("color:#cfe6ff;")
        left.addWidget(sep2)

        left.addWidget(section_title("Selected element"))

        left.addWidget(small_label("Font size (text only):"))
        self.txt_fontsize = QSpinBox()
        self.txt_fontsize.setRange(6, 80)
        self.txt_fontsize.valueChanged.connect(self._selected_text_font_changed)
        left.addWidget(self.txt_fontsize)

        info_align = QLabel("Alignment: center (locked)")
        info_align.setStyleSheet("font-weight:900; font-size:12px; color:#345b7a;")
        left.addWidget(info_align)

        left.addWidget(small_label("Move selected element (arrows):"))

        arrow_row1 = QHBoxLayout()
        self.btn_left = QPushButton()
        self.btn_right = QPushButton()
        self.btn_left.setIcon(self.style().standardIcon(QStyle.SP_ArrowLeft))
        self.btn_right.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
        for b in (self.btn_left, self.btn_right):
            b.setMinimumHeight(44)
            b.setMinimumWidth(120)
            b.setStyleSheet("""
                QPushButton { background:#2d7ef7; color:white; border-radius:12px; font-weight:900; }
                QPushButton:hover { background:#236be0; }
                QPushButton:pressed { background:#1c58ba; }
            """)
        self.btn_left.clicked.connect(lambda: self._nudge_selected(-2, 0))
        self.btn_right.clicked.connect(lambda: self._nudge_selected(2, 0))
        arrow_row1.addWidget(self.btn_left)
        arrow_row1.addWidget(self.btn_right)
        left.addLayout(arrow_row1)

        arrow_row2 = QHBoxLayout()
        self.btn_up = QPushButton()
        self.btn_down = QPushButton()
        self.btn_up.setIcon(self.style().standardIcon(QStyle.SP_ArrowUp))
        self.btn_down.setIcon(self.style().standardIcon(QStyle.SP_ArrowDown))
        for b in (self.btn_up, self.btn_down):
            b.setMinimumHeight(44)
            b.setMinimumWidth(120)
            b.setStyleSheet("""
                QPushButton { background:#2d7ef7; color:white; border-radius:12px; font-weight:900; }
                QPushButton:hover { background:#236be0; }
                QPushButton:pressed { background:#1c58ba; }
            """)
        self.btn_up.clicked.connect(lambda: self._nudge_selected(0, -2))
        self.btn_down.clicked.connect(lambda: self._nudge_selected(0, 2))
        arrow_row2.addWidget(self.btn_up)
        arrow_row2.addWidget(self.btn_down)
        left.addLayout(arrow_row2)

        self.btn_center_h = QPushButton("Center Horizontally")
        self.btn_center_h.setMinimumHeight(44)
        self.btn_center_h.setStyleSheet("""
            QPushButton { background:#2d7ef7; color:white; border-radius:12px; font-weight:900; }
            QPushButton:hover { background:#236be0; }
            QPushButton:pressed { background:#1c58ba; }
        """)
        self.btn_center_h.clicked.connect(self._center_selected_horizontal)
        left.addWidget(self.btn_center_h)

        sep3 = QFrame()
        sep3.setFrameShape(QFrame.HLine)
        sep3.setStyleSheet("color:#cfe6ff;")
        left.addWidget(sep3)

        left.addWidget(section_title("Barcode (always)"))

        left.addWidget(small_label("Module width:"))
        self.bc_module = QSpinBox()
        self.bc_module.setRange(1, 10)
        self.bc_module.valueChanged.connect(self._barcode_module_from_spin)
        left.addWidget(self.bc_module)

        left.addWidget(small_label("Height fraction:"))
        self.bc_height = QDoubleSpinBox()
        self.bc_height.setRange(0.06, 0.50)
        self.bc_height.setSingleStep(0.01)
        self.bc_height.valueChanged.connect(self._barcode_height_from_spin)
        left.addWidget(self.bc_height)

        left.addWidget(small_label("Width fraction:"))
        self.bc_widthfrac = QDoubleSpinBox()
        self.bc_widthfrac.setRange(0.10, 0.98)
        self.bc_widthfrac.setSingleStep(0.01)
        self.bc_widthfrac.valueChanged.connect(self._barcode_widthfrac_from_spin)
        left.addWidget(self.bc_widthfrac)

        left.addWidget(small_label("Quiet zone:"))
        self.bc_quiet = QDoubleSpinBox()
        self.bc_quiet.setRange(0.0, 20.0)
        self.bc_quiet.setSingleStep(0.5)
        self.bc_quiet.valueChanged.connect(self._barcode_quiet_from_spin)
        left.addWidget(self.bc_quiet)

        left.addStretch(1)

        self.scroll_left = QScrollArea()
        self.scroll_left.setWidgetResizable(True)
        self.scroll_left.setWidget(inspector_widget)
        self.scroll_left.setFixedWidth(360)
        root.addWidget(self.scroll_left)

        self.scene = QGraphicsScene()
        self.view = LabelClickView(self.scene)
        self.view.set_click_callback(self._handle_scene_click)
        root.addWidget(self.view, stretch=1)

        right = QVBoxLayout()
        right.addStretch(1)

        self.btn_save = QPushButton("Save")
        self.btn_save.setMinimumHeight(52)
        self.btn_save.setStyleSheet("background:#2d7ef7;color:white;font-weight:900;border-radius:12px;")
        self.btn_save.clicked.connect(self._save_only)
        right.addWidget(self.btn_save)

        self.btn_apply = QPushButton("Apply")
        self.btn_apply.setMinimumHeight(52)
        self.btn_apply.setStyleSheet("background:#19b36a;color:white;font-weight:900;border-radius:12px;")
        self.btn_apply.clicked.connect(self._apply_and_close)
        right.addWidget(self.btn_apply)

        right.addStretch(2)

        right_frame = QFrame()
        right_frame.setLayout(right)
        right_frame.setFixedWidth(140)
        right_frame.setStyleSheet("QFrame { background: transparent; }")
        root.addWidget(right_frame)

        self.setMinimumSize(1120, 700)
        self.resize(1300, 760)

        self._pix_item: Optional[QGraphicsPixmapItem] = None
        self._border_item: Optional[QGraphicsRectItem] = None

        self._build_scene_shell()
        self._load_barcode_controls()
        self._update_selected_controls_enabled()

    def showEvent(self, event):
        super().showEvent(event)
        if not self._first_show_done:
            self._first_show_done = True
            QTimer.singleShot(0, lambda: self.request_render(immediate=True))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._border_item:
            self.view.fitInView(self._border_item, Qt.KeepAspectRatio)

    def closeEvent(self, event):
        self._closing = True
        self._pending_timer.stop()
        self._stop_thread(wait=True)
        super().closeEvent(event)

    def _stop_thread(self, wait: bool = False):
        try:
            if self._thread and self._thread.isRunning():
                self._thread.quit()
                if wait:
                    self._thread.wait(1500)
        except Exception:
            pass
        if wait:
            self._thread = None
            self._worker = None
            self._render_in_flight = False
            self._rerender_queued = False

    def _build_scene_shell(self):
        self.scene.clear()

        w_dots = mm_to_dots(self.w_mm, self.settings.dpi)
        h_dots = mm_to_dots(self.h_mm, self.settings.dpi)

        placeholder = QImage(w_dots, h_dots, QImage.Format_Grayscale8)
        placeholder.fill(255)

        self._pix_item = QGraphicsPixmapItem(QPixmap.fromImage(placeholder))
        self._pix_item.setPos(0, 0)
        self.scene.addItem(self._pix_item)

        border = QGraphicsRectItem(0, 0, w_dots, h_dots)
        border.setPen(QPen(QColor(15, 15, 15), 3))
        border.setBrush(Qt.NoBrush)
        border.setZValue(10)
        self.scene.addItem(border)
        self._border_item = border

        self.scene.setSceneRect(-30, -30, w_dots + 60, h_dots + 60)
        self.view.fitInView(border, Qt.KeepAspectRatio)

    def _maybe_reload_for_size_change(self):
        new_w, new_h = self.settings.selected_label_size_mm
        if (new_w, new_h) != self._loaded_size:
            self.w_mm, self.h_mm = new_w, new_h
            self._loaded_size = (new_w, new_h)
            self.layout_data = json.loads(json.dumps(self.layout_store.get_layout(self.w_mm, self.h_mm)))
            self._render_cache.clear()
            self.selected_element = None
            self._last_bboxes = {}
            self._build_scene_shell()
            self._load_barcode_controls()
            self._update_selected_controls_enabled()
            self.font_combo.blockSignals(True)
            ff = str(self.layout_data.get("font_family", "Arial")) or "Arial"
            if ff in {"MT Extra", "Marlett", "Symbol", "Webdings"}:
                ff = "Arial"
                self.layout_data["font_family"] = "Arial"
            if self.font_combo.findText(ff) >= 0:
                self.font_combo.setCurrentText(ff)
            else:
                self.layout_data["font_family"] = "Arial"
                self.font_combo.setCurrentText("Arial")
            self.font_combo.blockSignals(False)

    def _make_request_key(self) -> str:
        code, desc, price = self.get_current_row()
        company = self.parent().company_edit.text().strip()
        payload = {
            "dpi": self.settings.dpi,
            "w_mm": self.w_mm,
            "h_mm": self.h_mm,
            "layout": self.layout_data,
            "barcode_type": self.settings.barcode_type,
            "company": company,
            "code": code,
            "desc": desc,
            "price": price
        }
        return sha1_text(json.dumps(payload, sort_keys=True, ensure_ascii=False))

    def request_render(self, immediate: bool = False):
        if self._closing:
            return
        if immediate:
            self._pending_timer.stop()
            self._start_render()
        else:
            self._pending_timer.start(80)

    def _start_render(self):
        if self._closing:
            return
        self._maybe_reload_for_size_change()

        req_key = self._make_request_key()
        self._active_request_key = req_key

        # cache is safe, but we explicitly clear it on any position changes to avoid “not moving” issues
        if req_key in self._render_cache:
            self._apply_render_result(self._render_cache[req_key])
            return

        if self._render_in_flight:
            self._rerender_queued = True
            return

        code, desc, price = self.get_current_row()
        company = self.parent().company_edit.text().strip()

        self._render_in_flight = True
        self._rerender_queued = False

        self._thread = QThread(self)
        self._worker = RenderWorker(
            renderer=self.renderer,
            dpi=self.settings.dpi,
            w_mm=self.w_mm,
            h_mm=self.h_mm,
            layout=self.layout_data,
            barcode_type=self.settings.barcode_type,
            company=company,
            code=code,
            desc=desc,
            price=price,
            request_key=req_key
        )
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.finished.connect(self._on_render_finished)
        self._worker.failed.connect(self._on_render_failed)

        self._worker.finished.connect(self._thread.quit)
        self._worker.failed.connect(self._thread.quit)

        self._thread.finished.connect(self._thread.deleteLater)
        self._worker.finished.connect(self._worker.deleteLater)
        self._worker.failed.connect(self._worker.deleteLater)

        self._thread.start()

    def _on_render_failed(self, msg: str):
        if self._closing:
            return
        self._render_in_flight = False
        self._update_issues_ui([f"Render error: {msg}"])

        if self._rerender_queued and not self._closing:
            self._rerender_queued = False
            QTimer.singleShot(0, self._start_render)

    def _on_render_finished(self, result: RenderResult):
        if self._closing:
            return
        self._render_cache[result.request_key] = result
        self._render_in_flight = False

        if result.request_key == self._active_request_key:
            self._apply_render_result(result)

        if self._rerender_queued and not self._closing:
            self._rerender_queued = False
            QTimer.singleShot(0, self._start_render)

    def _apply_render_result(self, result: RenderResult):
        if self._pix_item:
            pix = QPixmap.fromImage(result.qimage)
            self._pix_item.setPixmap(pix)
        self._last_bboxes = result.bboxes or {}
        self._update_issues_ui(result.issues)

    def _update_issues_ui(self, issues: List[str]):
        if issues:
            msg = "\n".join(issues[:6]) + ("" if len(issues) <= 6 else "\n...")
            self.lbl_issues.setText(msg)
            self.lbl_issues.setStyleSheet("font-weight:900; font-size:13px; color:#d52828;")
        else:
            self.lbl_issues.setText("No issues ✓")
            self.lbl_issues.setStyleSheet("font-weight:900; font-size:13px; color:#118a3b;")

    def _handle_scene_click(self, scene_pos):
        if not self._pix_item:
            return
        pr = self._pix_item.boundingRect()
        if not pr.contains(scene_pos):
            return

        x = int(scene_pos.x())
        y = int(scene_pos.y())

        hits = []
        for name, (bx, by, bw, bh) in (self._last_bboxes or {}).items():
            if bx <= x <= (bx + bw) and by <= y <= (by + bh):
                area = max(1, bw * bh)
                hits.append((area, name))
        if not hits:
            return

        hits.sort(key=lambda t: t[0])
        chosen = hits[0][1]
        self._set_selected_element(chosen)

    def _set_selected_element(self, name: str):
        self.selected_element = name
        self.lbl_selected.setText(f"Selected: {name}")
        self._load_selected_into_controls()
        self._update_selected_controls_enabled()

    def _update_selected_controls_enabled(self):
        text_ok = self.selected_element in ("Description", "Price", "Barcode Text", "Company Name")
        self.txt_fontsize.setEnabled(bool(text_ok))
        arrows_ok = self.selected_element is not None
        for b in (self.btn_left, self.btn_right, self.btn_up, self.btn_down):
            b.setEnabled(arrows_ok)
        self.btn_center_h.setEnabled(arrows_ok)

    def _load_selected_into_controls(self):
        key_map = {
            "Description": "description",
            "Price": "price",
            "Barcode Text": "barcode_text",
            "Company Name": "company"
        }
        if self.selected_element in key_map:
            el = self.layout_data.get("elements", {}).get(key_map[self.selected_element], {})
            self.txt_fontsize.blockSignals(True)
            self.txt_fontsize.setValue(int(el.get("font", 14)))
            self.txt_fontsize.blockSignals(False)

    def _label_dots(self) -> Tuple[int, int]:
        return mm_to_dots(self.w_mm, self.settings.dpi), mm_to_dots(self.h_mm, self.settings.dpi)

    def _nudge_selected(self, dx_dots: int, dy_dots: int):
        if not self.selected_element:
            return

        w_dots, h_dots = self._label_dots()
        if w_dots <= 0 or h_dots <= 0:
            return

        els = self.layout_data.get("elements", {})

        if self.selected_element == "Barcode":
            bc = els.get("barcode", {})
            bc["x"] = float(clamp(float(bc.get("x", 0.07)) + (dx_dots / w_dots), 0.0, 1.0))
            bc["y"] = float(clamp(float(bc.get("y", 0.68)) + (dy_dots / h_dots), 0.0, 1.0))
            els["barcode"] = bc
        else:
            key_map = {
                "Description": "description",
                "Price": "price",
                "Barcode Text": "barcode_text",
                "Company Name": "company"
            }
            key = key_map.get(self.selected_element)
            if not key or key not in els:
                return
            el = els.get(key, {})
            el["align"] = "center"

            # force float math + avoid any accidental string values
            cur_x = float(el.get("x", 0.5))
            cur_y = float(el.get("y", 0.5))
            el["x"] = float(clamp(cur_x + (float(dx_dots) / float(w_dots)), 0.0, 1.0))
            el["y"] = float(clamp(cur_y + (float(dy_dots) / float(h_dots)), 0.0, 1.0))
            els[key] = el

        self.layout_data["elements"] = els

        # IMPORTANT: clear cache so movement always shows immediately
        self._render_cache.clear()

        self.request_render(immediate=True)

    def _center_selected_horizontal(self):
        if not self.selected_element:
            return

        els = self.layout_data.get("elements", {})

        if self.selected_element == "Barcode":
            bc = els.get("barcode", {})
            width_frac = float(bc.get("width_frac", 0.78))
            width_frac = float(clamp(width_frac, 0.10, 0.98))
            bc["x"] = float(clamp((1.0 - width_frac) / 2.0, 0.0, 1.0))
            els["barcode"] = bc
        else:
            key_map = {
                "Description": "description",
                "Price": "price",
                "Barcode Text": "barcode_text",
                "Company Name": "company"
            }
            key = key_map.get(self.selected_element)
            if not key or key not in els:
                return
            el = els[key]
            el["align"] = "center"
            el["x"] = 0.50
            els[key] = el

        self.layout_data["elements"] = els
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _autofit_changed(self):
        self.layout_data["auto_fit"] = self.chk_autofit.isChecked()
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _font_changed(self, family: str):
        banned = {"MT Extra", "Marlett", "Symbol", "Webdings"}
        fam = (family or "").strip()
        if not fam or fam in banned:
            fam = "Arial"
        try:
            if not (self.renderer._find_font_path(fam, False) or self.renderer._find_font_path(fam, True)):
                fam = "Arial"
        except Exception:
            fam = "Arial"

        self.layout_data["font_family"] = str(fam)
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _selected_text_font_changed(self, v: int):
        if not self.selected_element:
            return
        key_map = {
            "Description": "description",
            "Price": "price",
            "Barcode Text": "barcode_text",
            "Company Name": "company"
        }
        key = key_map.get(self.selected_element)
        if not key:
            return
        el = self.layout_data.get("elements", {}).get(key, {})
        el["font"] = int(v)
        el["align"] = "center"
        self.layout_data["elements"][key] = el
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _load_barcode_controls(self):
        bc = self.layout_data.get("elements", {}).get("barcode", {})
        self.bc_module.blockSignals(True)
        self.bc_height.blockSignals(True)
        self.bc_widthfrac.blockSignals(True)
        self.bc_quiet.blockSignals(True)

        self.bc_module.setValue(int(round(float(bc.get("module_w", 2.0)))))
        self.bc_height.setValue(float(bc.get("bar_h_frac", 0.18)))
        self.bc_widthfrac.setValue(float(bc.get("width_frac", 0.78)))
        self.bc_quiet.setValue(float(bc.get("quiet_zone", 3.0)))

        self.bc_module.blockSignals(False)
        self.bc_height.blockSignals(False)
        self.bc_widthfrac.blockSignals(False)
        self.bc_quiet.blockSignals(False)

    def _barcode_module_from_spin(self, v: int):
        bc = self.layout_data.get("elements", {}).get("barcode", {})
        bc["module_w"] = float(max(1, int(v)))
        self.layout_data["elements"]["barcode"] = bc
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _barcode_height_from_spin(self, v: float):
        bc = self.layout_data.get("elements", {}).get("barcode", {})
        bc["bar_h_frac"] = float(clamp(float(v), 0.06, 0.50))
        self.layout_data["elements"]["barcode"] = bc
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _barcode_widthfrac_from_spin(self, v: float):
        bc = self.layout_data.get("elements", {}).get("barcode", {})
        bc["width_frac"] = float(clamp(float(v), 0.10, 0.98))
        self.layout_data["elements"]["barcode"] = bc
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _barcode_quiet_from_spin(self, v: float):
        bc = self.layout_data.get("elements", {}).get("barcode", {})
        bc["quiet_zone"] = float(clamp(float(v), 0.0, 20.0))
        self.layout_data["elements"]["barcode"] = bc
        self._render_cache.clear()
        self.request_render(immediate=True)

    def _normalize_layout_before_save(self):
        self.layout_data.setdefault("font_family", "Arial")
        els = self.layout_data.get("elements", {})
        for k in ("description", "price", "barcode_text", "company"):
            if k in els:
                els[k]["align"] = "center"
        bc = els.get("barcode", {})
        bc.setdefault("width_frac", 0.78)
        bc.setdefault("quiet_zone", 3.0)
        els["barcode"] = bc
        self.layout_data["elements"] = els

    def _save_only(self):
        self._normalize_layout_before_save()
        self.layout_store.set_layout(self.w_mm, self.h_mm, self.layout_data)
        QMessageBox.information(self, "Saved", "Layout saved for this label size.")

    def _apply_and_close(self):
        self._normalize_layout_before_save()

        code, desc, price = self.get_current_row()
        company = self.parent().company_edit.text().strip()

        _pil_l, _bboxes, issues = self.renderer.render(
            dpi=self.settings.dpi,
            w_mm=self.w_mm,
            h_mm=self.h_mm,
            layout=self.layout_data,
            barcode_type=self.settings.barcode_type,
            company=company,
            item_code=code,
            desc=desc,
            price=price
        )

        blocking = [x for x in issues if x.startswith("Clipping:") or x.startswith("Overlap:")]
        if blocking:
            QMessageBox.warning(self, "Layout issues", "Fix these before applying:\n\n" + "\n".join(blocking))
            return

        self.layout_store.set_layout(self.w_mm, self.h_mm, self.layout_data)
        self.accept()

    def external_refresh(self):
        self.request_render(immediate=True)


# ----------------------------
# Background jobs (Excel / Print)
# ----------------------------
class ExcelImportSignals(QObject):
    progress = Signal(int)
    finished = Signal(list)
    failed = Signal(str)
    cancelled = Signal()


class ExcelImportJob(QRunnable):
    def __init__(self, path: str):
        super().__init__()
        self.path = path
        self.signals = ExcelImportSignals()
        self._cancel = False

    def cancel(self):
        self._cancel = True

    @Slot()
    def run(self):
        try:
            wb = load_workbook(self.path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)

            try:
                first = next(rows_iter)
            except StopIteration:
                self.signals.finished.emit([])
                return

            header = [str(x).strip().lower() if x is not None else "" for x in first]

            def find_col(names):
                for n in names:
                    if n in header:
                        return header.index(n)
                return None

            code_i = find_col(["item lookup code", "itemlookupcode", "code", "item code", "lookup code"])
            desc_i = find_col(["description", "desc", "name"])
            price_i = find_col(["price", "cost"])

            if code_i is None and desc_i is None and price_i is None:
                code_i, desc_i, price_i = 0, 1, 2
                treat_first_as_data = True
            else:
                if code_i is None:
                    code_i = 0
                if desc_i is None:
                    desc_i = 1
                if price_i is None:
                    price_i = 2
                treat_first_as_data = False

            def take_from(r):
                code = str(r[code_i]).strip() if code_i < len(r) and r[code_i] is not None else ""
                desc = str(r[desc_i]).strip() if desc_i < len(r) and r[desc_i] is not None else ""
                price = str(r[price_i]).strip() if price_i < len(r) and r[price_i] is not None else ""
                return code, desc, price

            rows: List[Tuple[str, str, str]] = []

            try:
                max_row = int(ws.max_row or 1)
            except Exception:
                max_row = 1
            total = max(1, max_row)

            done = 1

            if treat_first_as_data:
                code, desc, price = take_from(first)
                if code or desc or price:
                    rows.append((code, desc, price))

            for r in rows_iter:
                if self._cancel:
                    self.signals.cancelled.emit()
                    return
                code, desc, price = take_from(r)
                if code or desc or price:
                    rows.append((code, desc, price))
                done += 1
                if done % 200 == 0:
                    self.signals.progress.emit(int((done / total) * 100))

            self.signals.progress.emit(100)
            self.signals.finished.emit(rows)
        except Exception as e:
            self.signals.failed.emit(str(e))


class PrintSignals(QObject):
    progress = Signal(int)
    finished = Signal(int)
    failed = Signal(str)
    cancelled = Signal()


class PrintJob(QRunnable):
    def __init__(
        self,
        printer_name: str,
        dpi: int,
        w_mm: float,
        h_mm: float,
        layout: Dict[str, Any],
        barcode_type: str,
        company: str,
        rows: Optional[List[Tuple[str, str, str]]],
        doc_name: str,
        rows_file: Optional[str] = None,
        indeterminate: bool = False
    ):
        super().__init__()
        self.printer_name = printer_name
        self.dpi = dpi
        self.w_mm = w_mm
        self.h_mm = h_mm
        self.layout = layout
        self.barcode_type = barcode_type
        self.company = company
        self.rows = rows or []
        self.doc_name = doc_name
        self.rows_file = rows_file
        self.indeterminate = indeterminate

        self.signals = PrintSignals()
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def _iter_rows(self):
        if self.rows_file:
            try:
                with open(self.rows_file, "r", encoding="utf-8", errors="ignore") as f:
                    for ln in f:
                        if self._cancel:
                            return
                        s = (ln or "").strip()
                        if not s:
                            continue
                        parsed = parse_csv_line(s)
                        if not parsed:
                            continue
                        code, desc, price = parsed
                        code = (code or "").strip()
                        desc = (desc or "").strip()
                        price = (price or "").strip()
                        if not desc or not price:
                            continue
                        yield code, desc, price
            except Exception:
                return
        else:
            for r in self.rows:
                if self._cancel:
                    return
                yield r

    @Slot()
    def run(self):
        hPrinter = None
        try:
            if not self.printer_name:
                self.signals.failed.emit("No printer selected.")
                return
            if not PIL_AVAILABLE:
                self.signals.failed.emit("PIL is required to print bitmaps.")
                return

            renderer = BitmapLabelRenderer()
            w_dots = mm_to_dots(self.w_mm, self.dpi)
            h_dots = mm_to_dots(self.h_mm, self.dpi)

            hPrinter = win32print.OpenPrinter(self.printer_name)
            hJob = win32print.StartDocPrinter(hPrinter, 1, (self.doc_name, None, "RAW"))
            try:
                win32print.StartPagePrinter(hPrinter)

                printed = 0
                total_known = (len(self.rows) if (not self.rows_file) else 0)

                for (code, desc, price) in self._iter_rows():
                    if self._cancel:
                        self.signals.cancelled.emit()
                        return

                    pil_l, _b, _issues = renderer.render(
                        dpi=self.dpi,
                        w_mm=self.w_mm,
                        h_mm=self.h_mm,
                        layout=self.layout,
                        barcode_type=self.barcode_type,
                        company=self.company,
                        item_code=code,
                        desc=desc,
                        price=price
                    )

                    data, w, h, bpr = pil_to_1bpp_bytes(pil_l)
                    gfa = zpl_gfa_from_1bpp(data, w, h, bpr)
                    zpl = zpl_for_bitmap_label(w_dots, h_dots, gfa) + "\n"
                    win32print.WritePrinter(hPrinter, zpl.encode("utf-8", errors="ignore"))

                    printed += 1
                    if not self.indeterminate and total_known > 0 and printed % 5 == 0:
                        self.signals.progress.emit(int((printed / total_known) * 100))
                    elif self.indeterminate and printed % 50 == 0:
                        self.signals.progress.emit(0)

                if self._cancel:
                    self.signals.cancelled.emit()
                    return

                self.signals.progress.emit(100)
                self.signals.finished.emit(printed)

            finally:
                try:
                    win32print.EndPagePrinter(hPrinter)
                except Exception:
                    pass
                try:
                    win32print.EndDocPrinter(hPrinter)
                except Exception:
                    pass

        except Exception as e:
            self.signals.failed.emit(str(e))
        finally:
            if hPrinter:
                try:
                    win32print.ClosePrinter(hPrinter)
                except Exception:
                    pass
            if self.rows_file:
                try:
                    os.remove(self.rows_file)
                except Exception:
                    pass


# ----------------------------
# Main window
# ----------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        if os.path.exists(resource_path("icon.ico")):
            self.setWindowIcon(QIcon(resource_path("icon.ico")))

        self.settings = SettingsStore(SETTINGS_FILE)
        self.settings.load()
        self.settings.save()

        self.layout_store = LayoutStore(LAYOUT_FILE)
        self.renderer = BitmapLabelRenderer()
        self.preview_dialog: Optional[PreviewLayoutDialog] = None

        self.thread_pool = QThreadPool.globalInstance()
        self._progress_dialog: Optional[QProgressDialog] = None
        self._active_print_jobs: List[PrintJob] = []

        # Import safety (keep job reference + chunk safe append)
        self._excel_job: Optional[ExcelImportJob] = None
        self._fill_timer: Optional[QTimer] = None
        self._fill_lines: List[str] = []
        self._fill_index: int = 0
        self._fill_progress: Optional[QProgressDialog] = None
        self._fill_cursor: Optional[QTextCursor] = None
        self._fill_has_any: bool = False

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        header = QHBoxLayout()
        logo = QLabel()
        if os.path.exists(resource_path("icon.png")):
            logo.setPixmap(QPixmap(resource_path("icon.png")).scaled(26, 26, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        header.addWidget(logo)
        title = QLabel(APP_TITLE)
        title.setStyleSheet("font-size:24px;font-weight:900;")
        header.addWidget(title)
        header.addStretch(1)
        root.addLayout(header)

        topbar = QHBoxLayout()
        topbar.setSpacing(12)

        topbar.addWidget(QLabel("Printer:"))
        self.printer_combo = QComboBox()
        self.printer_combo.setMinimumWidth(380)
        self.printer_combo.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        topbar.addWidget(self.printer_combo)

        self.btn_refresh_printers = QPushButton("Refresh printers")
        self.btn_use_printer = QPushButton("Use selected printer")
        self.btn_refresh_printers.clicked.connect(self.refresh_printers)
        self.btn_use_printer.clicked.connect(self.use_selected_printer)
        topbar.addWidget(self.btn_refresh_printers)
        topbar.addWidget(self.btn_use_printer)

        topbar.addWidget(QLabel("Label size:"))
        self.size_combo = QComboBox()
        self.size_combo.setMinimumWidth(170)
        topbar.addWidget(self.size_combo)

        topbar.addWidget(QLabel("Unit:"))
        self.unit_combo = QComboBox()
        self.unit_combo.addItems(["mm", "inches"])
        self.unit_combo.setCurrentText(self.settings.unit)
        self.unit_combo.setMinimumWidth(110)
        topbar.addWidget(self.unit_combo)

        self.btn_add_size = QPushButton("Add size")
        self.btn_remove_size = QPushButton("Remove size")
        self.btn_add_size.clicked.connect(self.add_size)
        self.btn_remove_size.clicked.connect(self.remove_size)
        topbar.addWidget(self.btn_add_size)
        topbar.addWidget(self.btn_remove_size)

        topbar.addWidget(QLabel("Barcode:"))
        self.barcode_combo = QComboBox()
        self.barcode_combo.addItems(["Code128", "Code39"])
        self.barcode_combo.setCurrentText(self.settings.barcode_type)
        self.barcode_combo.setMinimumWidth(120)
        topbar.addWidget(self.barcode_combo)

        topbar.addStretch(1)
        root.addLayout(topbar)

        cbox = QHBoxLayout()
        cbox.addWidget(QLabel("Company name:"))

        self.company_edit = QLineEdit(self.settings.data.get("company_name", "") or "")
        self.company_edit.setPlaceholderText("Type company name")
        cbox.addWidget(self.company_edit)

        self.btn_save_company = QPushButton("Save")
        self.btn_save_company.setMinimumWidth(90)
        self.btn_save_company.clicked.connect(self.save_company_name)
        cbox.addWidget(self.btn_save_company)

        root.addLayout(cbox)

        self.format_hint = QLabel(
        "Format:\n"
        "code, description, price\n"
        "description, price (if there's no code)"
        )
        self.format_hint.setWordWrap(True)
        self.format_hint.setStyleSheet("font-weight:900; font-size:14px; color:#13324f;")
        root.addWidget(self.format_hint)


        self.lines_edit = QTextEdit()
        self.lines_edit.setMinimumHeight(360)
        self.lines_edit.setPlaceholderText("")
        root.addWidget(self.lines_edit, stretch=1)

        bottom = QHBoxLayout()
        bottom.setSpacing(12)

        self.btn_import_excel = QPushButton("Import Excel")
        self.btn_validate = QPushButton("Validate Lines")
        self.btn_preview = QPushButton("Preview Layout")
        self.btn_print_selected = QPushButton("Print Selected Row")
        self.btn_print_all = QPushButton("Print (All)")

        self.btn_import_excel.clicked.connect(self.import_excel)
        self.btn_validate.clicked.connect(self.validate_lines)
        self.btn_preview.clicked.connect(self.open_preview)
        self.btn_print_selected.clicked.connect(self.print_selected_row)
        self.btn_print_all.clicked.connect(self.print_all_rows)

        bottom.addWidget(self.btn_import_excel)
        bottom.addWidget(self.btn_validate)
        bottom.addStretch(1)
        bottom.addWidget(self.btn_preview)
        bottom.addWidget(self.btn_print_selected)
        bottom.addWidget(self.btn_print_all)
        root.addLayout(bottom)

        self.size_combo.currentTextChanged.connect(self.on_size_changed)
        self.unit_combo.currentTextChanged.connect(self.on_unit_changed)
        self.barcode_combo.currentTextChanged.connect(self.on_barcode_type_changed)
        self.company_edit.textChanged.connect(self.refresh_preview_if_open)
        self.lines_edit.textChanged.connect(self.refresh_preview_if_open)
        self.lines_edit.cursorPositionChanged.connect(self.refresh_preview_if_open)

        self.apply_styles()
        self.refresh_printers()
        self.update_size_combo()

        if not PIL_AVAILABLE:
            QMessageBox.warning(self, "Missing dependency", "PIL (Pillow) is required for bitmap preview/print.\nInstall: pip install pillow")
        if not BARCODE_AVAILABLE:
            QMessageBox.warning(self, "Missing dependency", "python-barcode is required for barcodes.\nInstall: pip install python-barcode[images]")

    def save_company_name(self):
        name = self.company_edit.text().strip()
        self.settings.data["company_name"] = name
        self.settings.save()
        QMessageBox.information(self, "Saved", "Company name saved.")

    def apply_styles(self):
        qss = """
        QMainWindow { background: #f3f9ff; }
        QLabel { color: #13324f; font-size: 13px; }

        QLineEdit, QComboBox, QTextEdit, QSpinBox, QDoubleSpinBox {
            background: white;
            color: #0f2b46;
            border: 1px solid #cfe6ff;
            border-radius: 10px;
            padding: 6px;
            selection-background-color: #2d7ef7;
            selection-color: white;
            font-size: 13px;
        }

        QComboBox::drop-down { border: 0px; width: 28px; }
        QComboBox QAbstractItemView {
            background: white;
            color: #0f2b46;
            selection-background-color: #2d7ef7;
            selection-color: white;
            border: 1px solid #cfe6ff;
        }

        QPushButton {
            background: #2d7ef7;
            color: white;
            border: none;
            border-radius: 10px;
            padding: 10px 14px;
            font-weight: 900;
        }
        QPushButton:hover { background: #236be0; }
        QPushButton:pressed { background: #1c58ba; }

        QCheckBox { color: #0f2b46; font-weight: 900; font-size: 13px; }
        QCheckBox::indicator { width: 18px; height: 18px; }

        QMessageBox { background: white; }
        QMessageBox QLabel { color: #0f2b46; font-size: 14px; }
        QMessageBox QPushButton { min-width: 110px; }

        QProgressDialog { background: white; }
        QProgressDialog QLabel { color: #0f2b46; font-size: 14px; }
        """
        self.setStyleSheet(qss)

    def _show_progress(self, title: str, text: str, indeterminate: bool = False):
        if self._progress_dialog:
            try:
                self._progress_dialog.close()
                self._progress_dialog.deleteLater()
            except Exception:
                pass
            self._progress_dialog = None

        if indeterminate:
            dlg = QProgressDialog(text, "Cancel", 0, 0, self)
        else:
            dlg = QProgressDialog(text, "Cancel", 0, 100, self)

        dlg.setWindowTitle(title)
        dlg.setWindowModality(Qt.WindowModal)
        dlg.setAutoClose(False)
        dlg.setAutoReset(False)
        dlg.setMinimumDuration(0)
        if not indeterminate:
            dlg.setValue(0)
        self._progress_dialog = dlg
        dlg.show()
        return dlg

    def _close_progress(self):
        if self._progress_dialog:
            dlg = self._progress_dialog
            self._progress_dialog = None
            try:
                dlg.reset()
                dlg.hide()
                dlg.close()
                dlg.deleteLater()
            except Exception:
                pass

    def refresh_printers(self):
        self.printer_combo.clear()
        try:
            printers = [p[2] for p in win32print.EnumPrinters(2)]
        except Exception as e:
            printers = []
            QMessageBox.warning(self, "Printer Error", f"Could not list printers:\n{e}")

        self.printer_combo.addItems(printers)
        if printers:
            if self.settings.selected_printer and self.settings.selected_printer in printers:
                self.printer_combo.setCurrentText(self.settings.selected_printer)
            else:
                self.printer_combo.setCurrentIndex(0)

    def use_selected_printer(self):
        pr = self.printer_combo.currentText().strip()
        self.settings.data["selected_printer"] = pr
        self.settings.save()
        QMessageBox.information(self, "Printer selected", f"Selected printer:\n{pr}")

    def update_size_combo(self):
        self.size_combo.blockSignals(True)
        self.size_combo.clear()

        unit = self.unit_combo.currentText()
        sizes = self.settings.label_sizes_mm
        if not sizes:
            sizes = [[76.2, 25.4]]
            self.settings.data["label_sizes_mm"] = sizes
            self.settings.data["selected_label_size_mm"] = sizes[0]
            self.settings.save()

        for w, h in sizes:
            if unit == "mm":
                txt = f"{w:.1f} x {h:.1f} mm"
            else:
                txt = f"{mm_to_inches(w):.2f} x {mm_to_inches(h):.2f} in"
            self.size_combo.addItem(txt)

        sw, sh = self.settings.selected_label_size_mm
        idx = 0
        for i, (w, h) in enumerate(sizes):
            if abs(w - sw) < 0.01 and abs(h - sh) < 0.01:
                idx = i
                break
        self.size_combo.setCurrentIndex(idx)
        self.size_combo.blockSignals(False)

    def on_unit_changed(self):
        self.settings.data["unit"] = self.unit_combo.currentText()
        self.settings.save()
        self.update_size_combo()
        self.refresh_preview_if_open()

    def on_size_changed(self):
        idx = self.size_combo.currentIndex()
        if idx < 0:
            return
        w_mm, h_mm = self.settings.label_sizes_mm[idx]
        self.settings.data["selected_label_size_mm"] = [float(w_mm), float(h_mm)]
        self.settings.save()
        self.refresh_preview_if_open()

    def add_size(self):
        unit = self.unit_combo.currentText()
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Label Size")
        form = QFormLayout(dialog)
        wspin = QDoubleSpinBox()
        hspin = QDoubleSpinBox()
        wspin.setMinimum(0.01)
        hspin.setMinimum(0.01)
        form.addRow(f"Width ({'mm' if unit=='mm' else 'in'}):", wspin)
        form.addRow(f"Height ({'mm' if unit=='mm' else 'in'}):", hspin)
        btns = QHBoxLayout()
        ok = QPushButton("OK")
        cancel = QPushButton("Cancel")
        ok.clicked.connect(dialog.accept)
        cancel.clicked.connect(dialog.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        form.addRow(btns)
        if dialog.exec() == QDialog.Accepted:
            w = float(wspin.value())
            h = float(hspin.value())
            if unit == "inches":
                w = inches_to_mm(w)
                h = inches_to_mm(h)
            self.settings.data["label_sizes_mm"].append([w, h])
            self.settings.save()
            self.update_size_combo()
            self.refresh_preview_if_open()

    def remove_size(self):
        idx = self.size_combo.currentIndex()
        sizes = self.settings.label_sizes_mm
        if idx < 0 or not sizes:
            return
        if len(sizes) <= 1:
            QMessageBox.warning(self, "Cannot remove", "At least one label size must remain.")
            return
        del sizes[idx]
        self.settings.data["label_sizes_mm"] = sizes
        new_idx = max(0, idx - 1)
        self.settings.data["selected_label_size_mm"] = sizes[new_idx]
        self.settings.save()
        self.update_size_combo()
        self.refresh_preview_if_open()

    def _get_nonempty_lines(self) -> List[str]:
        raw = self.lines_edit.toPlainText()
        lines = []
        for ln in raw.splitlines():
            s = (ln or "").strip()
            if s:
                lines.append(s)
        return lines

    def _parse_rows_from_editor(self) -> Tuple[List[Tuple[str, str, str]], List[str]]:
        rows: List[Tuple[str, str, str]] = []
        errors: List[str] = []

        lines = self._get_nonempty_lines()
        for i, ln in enumerate(lines, start=1):
            parsed = parse_csv_line(ln)
            if not parsed:
                errors.append(f"Line {i}: Invalid format (need 3 comma-separated values)")
                continue

            code, desc, price = parsed
            code = (code or "").strip()
            desc = (desc or "").strip()
            price = (price or "").strip()

            if not desc:
                errors.append(f"Line {i}: Empty description")
            if not price:
                errors.append(f"Line {i}: Empty price")
            else:
                p = price.replace("$", "").strip()
                try:
                    float(p)
                except Exception:
                    errors.append(f"Line {i}: Price is not numeric")

            if desc and price:
                rows.append((code, desc, price))

        return rows, errors

    def _current_line_text(self) -> str:
        cur = self.lines_edit.textCursor()
        return (cur.block().text() or "").strip()

    def get_current_row_data(self) -> Tuple[str, str, str]:
        ln = self._current_line_text()
        parsed = parse_csv_line(ln)
        if not parsed:
            return ("", "", "")
        code, desc, price = parsed
        return (code.strip(), desc.strip(), price.strip())

    def import_excel(self):
        start_dir = self.settings.data.get("last_excel_dir", "") or ""
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", start_dir, "Excel Files (*.xlsx)")
        if not path:
            return

        self.settings.data["last_excel_dir"] = os.path.dirname(path)
        self.settings.save()

        dlg = self._show_progress("Importing Excel", "Reading Excel…")
        job = ExcelImportJob(path)
        self._excel_job = job  # keep reference so it cannot be GC'ed mid-run

        dlg.canceled.connect(job.cancel)
        job.signals.progress.connect(dlg.setValue)
        job.signals.cancelled.connect(lambda: (self._close_progress(), QMessageBox.information(self, "Excel", "Import cancelled.")))
        job.signals.failed.connect(lambda msg: (self._close_progress(), QMessageBox.critical(self, "Excel Error", msg)))
        job.signals.finished.connect(self._on_excel_finished)
        self.thread_pool.start(job)

    def _cancel_fill(self):
        try:
            if self._fill_timer:
                self._fill_timer.stop()
                self._fill_timer.deleteLater()
        except Exception:
            pass
        self._fill_timer = None

        try:
            if self._fill_progress:
                self._fill_progress.reset()
                self._fill_progress.hide()
                self._fill_progress.close()
                self._fill_progress.deleteLater()
        except Exception:
            pass
        self._fill_progress = None

        self._fill_lines = []
        self._fill_index = 0
        self._fill_cursor = None
        self._fill_has_any = False

    def _start_chunk_fill(self, out_lines: List[str]):
        self._cancel_fill()

        self._fill_lines = out_lines
        self._fill_index = 0
        self._fill_has_any = False

        self._fill_progress = QProgressDialog("Loading into editor…", "Cancel", 0, 100, self)
        self._fill_progress.setWindowTitle("Importing Excel")
        self._fill_progress.setWindowModality(Qt.WindowModal)
        self._fill_progress.setAutoClose(True)
        self._fill_progress.setAutoReset(True)
        self._fill_progress.setMinimumDuration(0)
        self._fill_progress.canceled.connect(self._cancel_fill)
        self._fill_progress.show()

        self.lines_edit.blockSignals(True)
        self.lines_edit.setPlainText("")
        self.lines_edit.blockSignals(False)

        self._fill_cursor = QTextCursor(self.lines_edit.document())
        self._fill_cursor.movePosition(QTextCursor.End)

        self._fill_timer = QTimer(self)
        self._fill_timer.setSingleShot(False)
        self._fill_timer.timeout.connect(self._fill_tick_safe)
        self._fill_timer.start(1)

    def _fill_tick_safe(self):
        try:
            self._fill_tick()
        except Exception as e:
            self._cancel_fill()
            QMessageBox.critical(self, "Import Error", f"Import failed while writing to editor:\n{e}")

    def _fill_tick(self):
        if not self._fill_lines or not self._fill_cursor:
            self._cancel_fill()
            return

        chunk = 600
        end = min(len(self._fill_lines), self._fill_index + chunk)
        part = "\n".join(self._fill_lines[self._fill_index:end])

        if self._fill_has_any:
            part = "\n" + part

        self.lines_edit.blockSignals(True)
        self._fill_cursor.movePosition(QTextCursor.End)
        self._fill_cursor.insertText(part)
        self.lines_edit.blockSignals(False)

        self._fill_has_any = True
        self._fill_index = end

        if self._fill_progress:
            pct = int((self._fill_index / max(1, len(self._fill_lines))) * 100)
            self._fill_progress.setValue(pct)

        if self._fill_index >= len(self._fill_lines):
            count = len(self._fill_lines)
            self._cancel_fill()
            QMessageBox.information(self, "Excel Imported", f"Imported {count} item(s).")
            self.refresh_preview_if_open()

    def _on_excel_finished(self, rows: list):
        try:
            self._close_progress()
            self._excel_job = None

            out_lines: List[str] = []
            for (code, desc, price) in rows:
                code = (str(code) if code is not None else "").strip()
                desc = (str(desc) if desc is not None else "").strip()
                price = (str(price) if price is not None else "").strip()

                if not code and not desc and not price:
                    continue

                if code:
                    out_lines.append(f"{code}, {desc}, {price}")
                else:
                    out_lines.append(f"{desc}, {price}")

            if len(out_lines) >= 800:
                self._start_chunk_fill(out_lines)
                return

            self.lines_edit.blockSignals(True)
            self.lines_edit.setPlainText("\n".join(out_lines))
            self.lines_edit.blockSignals(False)

            QMessageBox.information(self, "Excel Imported", f"Imported {len(out_lines)} item(s).")
            self.refresh_preview_if_open()
        except Exception as e:
            self._close_progress()
            self._excel_job = None
            QMessageBox.critical(self, "Excel Error", f"Import failed:\n{e}")

    def validate_lines(self):
        _rows, errors = self._parse_rows_from_editor()
        if errors:
            QMessageBox.warning(self, "Validation errors", "\n".join(errors[:80]) + ("" if len(errors) <= 80 else "\n..."))
        else:
            QMessageBox.information(self, "Validation", "No errors found.")

    def on_barcode_type_changed(self):
        self.settings.data["barcode_type"] = self.barcode_combo.currentText()
        self.settings.save()
        self.refresh_preview_if_open()

    def refresh_preview_if_open(self):
        if self.preview_dialog and self.preview_dialog.isVisible():
            self.preview_dialog.external_refresh()

    def open_preview(self):
        self.preview_dialog = PreviewLayoutDialog(self, self.settings, self.layout_store, self.renderer, self.get_current_row_data)
        self.preview_dialog.exec()

    def _write_editor_to_tempfile(self) -> str:
        fd, path = tempfile.mkstemp(prefix="label_print_", suffix=".txt")
        os.close(fd)
        doc = self.lines_edit.document()
        block = doc.firstBlock()
        with open(path, "w", encoding="utf-8", errors="ignore") as f:
            while block.isValid():
                s = (block.text() or "").strip()
                if s:
                    f.write(s + "\n")
                block = block.next()
        return path

    def _on_print_cancelled(self, job: PrintJob):
        self._close_progress()
        try:
            if job in self._active_print_jobs:
                self._active_print_jobs.remove(job)
        except Exception:
            pass
        QMessageBox.information(self, "Print", "Print cancelled.")

    def _on_print_failed(self, job: PrintJob, msg: str):
        self._close_progress()
        try:
            if job in self._active_print_jobs:
                self._active_print_jobs.remove(job)
        except Exception:
            pass
        QMessageBox.critical(self, "Print error", msg)

    def _on_print_finished(self, job: PrintJob, count: int):
        self._close_progress()
        try:
            if job in self._active_print_jobs:
                self._active_print_jobs.remove(job)
        except Exception:
            pass
        QMessageBox.information(self, "Printed", f"Successfully printed {count} label(s).")

    def _start_print_job(self, rows: Optional[List[Tuple[str, str, str]]], doc_name: str, rows_file: Optional[str] = None, indeterminate: bool = False):
        printer = (self.settings.selected_printer or "").strip()
        if not printer:
            QMessageBox.warning(self, "Printer", "Select a printer first (Use selected printer).")
            return

        if (not rows_file) and (not rows):
            QMessageBox.warning(self, "Print", "No rows to print.")
            return

        w_mm, h_mm = self.settings.selected_label_size_mm
        layout = self.layout_store.get_layout(w_mm, h_mm)
        company = self.company_edit.text().strip()
        btype = self.settings.barcode_type
        dpi = self.settings.dpi

        dlg = self._show_progress("Printing", "Rendering and sending labels…", indeterminate=indeterminate)

        job = PrintJob(
            printer_name=printer,
            dpi=dpi,
            w_mm=w_mm,
            h_mm=h_mm,
            layout=layout,
            barcode_type=btype,
            company=company,
            rows=rows,
            doc_name=doc_name,
            rows_file=rows_file,
            indeterminate=indeterminate
        )

        self._active_print_jobs.append(job)

        dlg.canceled.connect(job.cancel)
        if not indeterminate:
            job.signals.progress.connect(dlg.setValue)

        job.signals.cancelled.connect(lambda j=job: self._on_print_cancelled(j))
        job.signals.failed.connect(lambda msg, j=job: self._on_print_failed(j, msg))
        job.signals.finished.connect(lambda count, j=job: self._on_print_finished(j, count))

        self.thread_pool.start(job)

    def print_selected_row(self):
        ln = (self.lines_edit.textCursor().block().text() or "").strip()
        parsed = parse_csv_line(ln)
        if not parsed:
            QMessageBox.warning(self, "Print", "Put the cursor on a valid line (description and price).")
            return
        code, desc, price = parsed
        if not (desc or "").strip() or not (price or "").strip():
            QMessageBox.warning(self, "Print", "Put the cursor on a valid line (description and price).")
            return
        self._start_print_job([(code.strip(), desc.strip(), price.strip())], "Label Print (Selected Row)", rows_file=None, indeterminate=False)

    def print_all_rows(self):
        tmp_path = self._write_editor_to_tempfile()
        try:
            if os.path.getsize(tmp_path) <= 0:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
                QMessageBox.warning(self, "Print", "No rows to print.")
                return
        except Exception:
            pass

        self._start_print_job(rows=None, doc_name="Label Print (All)", rows_file=tmp_path, indeterminate=True)

    def print_labels(self):
        self.print_all_rows()


def main():
    _install_global_exception_hook()
    app = QApplication(sys.argv)
    win = MainWindow()
    win.resize(1600, 900)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
